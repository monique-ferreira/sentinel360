"""
bi_excel.py — Gera relatório BI como arquivo Excel (.xlsx) para o Sentinel360.
"""

from __future__ import annotations

import io
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from collections import Counter


# ── Paleta (fundo branco, texto escuro — compatível com Excel padrão) ─────────

ACCENT  = "217346"   # verde Excel
BLUE    = "1F6AA5"
RED     = "C0392B"
YELLOW  = "B8860B"
PURPLE  = "6C3483"
HEADER_BG = "2F5496"
HEADER_FG = "FFFFFF"
ALT_ROW   = "EBF1F8"


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color="000000", size=10) -> Font:
    return Font(name="Calibri", bold=bold, color=color, size=size)


def _border() -> Border:
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)


def _header_row(ws, row: int, values: list[str]):
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill  = _fill(HEADER_BG)
        c.font  = Font(name="Calibri", bold=True, color=HEADER_FG, size=10)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = _border()


def _data_row(ws, row: int, values: list, alt: bool = False):
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        if alt:
            c.fill = _fill(ALT_ROW)
        c.font  = _font()
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        c.border = _border()


def _set_col_widths(ws, widths: list[int]):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def generate(cloud_items: list[dict], scan_history: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    generated_at = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    # ── Sheet 1: Resumo / KPIs ────────────────────────────────────────────────
    ws_kpi = wb.active
    ws_kpi.title = "Resumo"

    _set_col_widths(ws_kpi, [28, 20, 20, 20, 20])

    t = ws_kpi["A1"]
    t.value = "SENTINEL360 — RELATÓRIO BI"
    t.font  = Font(name="Calibri", bold=True, color=ACCENT, size=16)
    t.fill  = _fill("FFFFFF")

    ws_kpi["A2"] = f"Gerado em: {generated_at}"
    ws_kpi["A2"].font = _font(color="666666", size=9)
    ws_kpi["A2"].fill = _fill("FFFFFF")

    total     = len(cloud_items)
    inativos  = sum(1 for i in cloud_items if i.get("inativo") == "SIM")
    com_risco = sum(1 for i in cloud_items if i.get("riscos") not in ("NENHUM", "", None))
    total_mb  = sum(float(i.get("tamanho_mb", 0) or 0) for i in cloud_items)

    kpi_data = [
        ("Total de arquivos",        total,             BLUE),
        ("Arquivos inativos",        inativos,          YELLOW),
        ("Com risco / sensíveis",    com_risco,         RED),
        ("Armazenamento total (MB)", round(total_mb, 2), ACCENT),
    ]

    for col, (label, value, color) in enumerate(kpi_data, 1):
        lc = ws_kpi.cell(row=4, column=col, value=label)
        lc.fill  = _fill(HEADER_BG)
        lc.font  = Font(name="Calibri", bold=True, color=HEADER_FG, size=9)
        lc.alignment = Alignment(horizontal="center")
        lc.border = _border()

        vc = ws_kpi.cell(row=5, column=col, value=value)
        vc.fill  = _fill("FFFFFF")
        vc.font  = Font(name="Calibri", bold=True, color=color, size=20)
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.border = _border()
        ws_kpi.row_dimensions[5].height = 40

    # Risk breakdown
    risk_counter: Counter = Counter()
    for item in cloud_items:
        for r in (item.get("riscos") or "").split(","):
            r = r.strip()
            if r and r not in ("NENHUM", ""):
                risk_counter[r] += 1

    h7a = ws_kpi.cell(row=7, column=1, value="Categoria de Risco")
    h7a.font = _font(bold=True, color=ACCENT, size=11)
    h7a.fill = _fill("FFFFFF")
    h7b = ws_kpi.cell(row=7, column=2, value="Quantidade")
    h7b.font = _font(bold=True, color=ACCENT, size=11)
    h7b.fill = _fill("FFFFFF")
    for i, (label, count) in enumerate(risk_counter.most_common(), 1):
        ca = ws_kpi.cell(row=7 + i, column=1, value=label)
        ca.font = _font()
        ca.fill = _fill("FFFFFF")
        cb = ws_kpi.cell(row=7 + i, column=2, value=count)
        cb.font = _font(bold=True, color=RED)
        cb.fill = _fill("FFFFFF")

    # ── Sheet 2: Todos os arquivos ────────────────────────────────────────────
    ws_files = wb.create_sheet("Arquivos")
    _set_col_widths(ws_files, [35, 50, 20, 25, 12, 20, 18])

    ws_files["A1"].value = "Resultados de Varredura"
    ws_files["A1"].font = Font(name="Calibri", bold=True, color=ACCENT, size=13)
    ws_files["A1"].fill = _fill("FFFFFF")

    _header_row(ws_files, 2, ["Nome", "Caminho / URL", "Origem", "Riscos", "Inativo", "Tamanho (MB)", "Último Scan"])

    for i, item in enumerate(cloud_items):
        _data_row(ws_files, i + 3, [
            item.get("nome", ""),
            item.get("caminho", ""),
            item.get("origem", ""),
            item.get("riscos", "NENHUM"),
            item.get("inativo", "NÃO"),
            item.get("tamanho_mb", 0),
            item.get("last_scan", ""),
        ], alt=i % 2 == 1)

        risk_cell = ws_files.cell(row=i + 3, column=4)
        if item.get("riscos") not in ("NENHUM", "", None):
            risk_cell.font = _font(color=RED, bold=True)
        if item.get("inativo") == "SIM":
            ws_files.cell(row=i + 3, column=5).font = _font(color=YELLOW, bold=True)

    # ── Sheet 3: Arquivos de risco ────────────────────────────────────────────
    ws_risk = wb.create_sheet("Risco")
    _set_col_widths(ws_risk, [35, 50, 20, 25, 12, 18])

    risky = [i for i in cloud_items if i.get("riscos") not in ("NENHUM", "", None)]
    ws_risk["A1"].value = f"Arquivos de Alto Risco ({len(risky)} itens)"
    ws_risk["A1"].font = Font(name="Calibri", bold=True, color=RED, size=13)
    ws_risk["A1"].fill = _fill("FFFFFF")

    _header_row(ws_risk, 2, ["Nome", "Caminho / URL", "Origem", "Riscos", "Inativo", "Último Scan"])

    for i, item in enumerate(risky):
        _data_row(ws_risk, i + 3, [
            item.get("nome", ""),
            item.get("caminho", ""),
            item.get("origem", ""),
            item.get("riscos", ""),
            item.get("inativo", ""),
            item.get("last_scan", ""),
        ], alt=i % 2 == 1)
        ws_risk.cell(row=i + 3, column=4).font = _font(color=RED, bold=True)
        if item.get("inativo") == "SIM":
            ws_risk.cell(row=i + 3, column=5).font = _font(color=YELLOW, bold=True)

    # ── Sheet 4: Histórico de scans ───────────────────────────────────────────
    ws_hist = wb.create_sheet("Histórico")
    _set_col_widths(ws_hist, [22, 20, 20, 14, 14])

    ws_hist["A1"].value = "Histórico de Varreduras"
    ws_hist["A1"].font = Font(name="Calibri", bold=True, color=ACCENT, size=13)
    ws_hist["A1"].fill = _fill("FFFFFF")

    _header_row(ws_hist, 2, ["Data / Hora", "Tipo", "Arquivos Analisados", "Com Risco", "Inativos"])

    for i, h in enumerate(scan_history[:50]):
        _data_row(ws_hist, i + 3, [
            h.get("data", ""),
            h.get("tipo", ""),
            h.get("total_arquivos", 0),
            h.get("com_risco", 0),
            h.get("inativos", 0),
        ], alt=i % 2 == 1)
        if h.get("com_risco", 0) > 0:
            ws_hist.cell(row=i + 3, column=4).font = _font(color=RED, bold=True)
        if h.get("inativos", 0) > 0:
            ws_hist.cell(row=i + 3, column=5).font = _font(color=YELLOW, bold=True)

    # ── Sheet 5: Gráficos ─────────────────────────────────────────────────────
    ws_charts = wb.create_sheet("Gráficos")
    ws_charts["A1"].value = "Gráficos de Análise"
    ws_charts["A1"].font = Font(name="Calibri", bold=True, color=ACCENT, size=13)
    ws_charts["A1"].fill = _fill("FFFFFF")

    # ── Dados para gráfico de pizza: categorias de risco (cols A-B a partir da linha 3)
    risk_list = list(risk_counter.most_common(8))
    pie_data_row_start = 3

    if risk_list:
        ws_charts.cell(row=pie_data_row_start - 1, column=1, value="Categoria").font = _font(bold=True)
        ws_charts.cell(row=pie_data_row_start - 1, column=2, value="Qtd").font = _font(bold=True)
        for r_i, (label, count) in enumerate(risk_list):
            ws_charts.cell(row=pie_data_row_start + r_i, column=1, value=label)
            ws_charts.cell(row=pie_data_row_start + r_i, column=2, value=count)

        pie = PieChart()
        pie.title  = "Categorias de Risco"
        pie.style  = 2
        pie.width  = 18
        pie.height = 14
        data_ref   = Reference(ws_charts, min_col=2, min_row=pie_data_row_start - 1,
                                max_row=pie_data_row_start + len(risk_list) - 1)
        labels_ref = Reference(ws_charts, min_col=1, min_row=pie_data_row_start,
                                max_row=pie_data_row_start + len(risk_list) - 1)
        pie.add_data(data_ref, titles_from_data=True)
        pie.set_categories(labels_ref)
        ws_charts.add_chart(pie, "D2")   # pizza: D2, ocupa ~24 linhas

    # ── Dados para gráfico de barras: histórico — cols A-C fixas abaixo dos dados da pizza
    # Posição fixa na linha 30 para garantir que não sobreponha a pizza (que ocupa D2:M25)
    BAR_DATA_ROW = 30
    hist_slice = list(reversed(scan_history[:10]))

    if hist_slice:
        ws_charts.cell(row=BAR_DATA_ROW - 1, column=1, value="Data").font = _font(bold=True)
        ws_charts.cell(row=BAR_DATA_ROW - 1, column=2, value="Total").font = _font(bold=True)
        ws_charts.cell(row=BAR_DATA_ROW - 1, column=3, value="Com Risco").font = _font(bold=True)
        for h_i, hh in enumerate(hist_slice):
            ws_charts.cell(row=BAR_DATA_ROW + h_i, column=1, value=hh.get("data", "")[:10])
            ws_charts.cell(row=BAR_DATA_ROW + h_i, column=2, value=hh.get("total_arquivos", 0))
            ws_charts.cell(row=BAR_DATA_ROW + h_i, column=3, value=hh.get("com_risco", 0))

        bar = BarChart()
        bar.type    = "col"
        bar.title   = "Histórico de Scans"
        bar.y_axis.title = "Arquivos"
        bar.x_axis.title = "Data"
        bar.style   = 2
        bar.width   = 22
        bar.height  = 14
        data_ref = Reference(ws_charts, min_col=2, max_col=3,
                             min_row=BAR_DATA_ROW - 1, max_row=BAR_DATA_ROW + len(hist_slice) - 1)
        cats_ref = Reference(ws_charts, min_col=1,
                             min_row=BAR_DATA_ROW, max_row=BAR_DATA_ROW + len(hist_slice) - 1)
        bar.add_data(data_ref, titles_from_data=True)
        bar.set_categories(cats_ref)
        ws_charts.add_chart(bar, "D28")  # barras: D28, logo abaixo da pizza

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
